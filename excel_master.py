import openpyxl
import pandas as pd
import os


def get_cell_coordinate(row, column):
    letters = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    letter = letters[column]
    number = str(row)
    return f'{letter}{number}'


def create_column(table, property_name_list):
    if not os.path.isfile(table):
        workbook = openpyxl.Workbook()
    else:
        workbook = openpyxl.load_workbook(table)
    sheet = workbook.active

    column = 0
    for value in property_name_list:
        sheet.cell(row=1, column=column + 1, value=value)
        sheet[1][column].value = value
        column += 1

    workbook.save(table)


def import_xl(table):
    workbook = openpyxl.load_workbook(table)
    sheet = workbook.active

    property_ = []

    for cell in sheet[1]:
        if cell.value is not None:
            property_.append(cell.value)
    workbook.save(table)

    return property_


def property_export(row, table, site_prop_list):
    workbook = openpyxl.load_workbook(table)
    sheet = workbook.active

    table_prop_list = import_xl(table)
    for prop in site_prop_list:
        if prop[0] in table_prop_list:
            sheet.cell(row=row, column=table_prop_list.index(prop[0]) + 1).value = prop[1]

    # img = Image(fr'C:\Users\User\PycharmProjects\Freelans_1\фото\{site_prop_list[0][1]}.jpg')
    # cell = get_cell_coordinate(row=row, column=table_prop_list.index('Фото'))
    # sheet.add_image(img, cell)
    # sheet.row_dimensions[row].height = 135

    workbook.save(table)
    return


def find_row(table, name):
    df = pd.read_excel(table)
    row_index = df.index[df['Наименование'] == name].tolist()

    if row_index:
        return row_index[0]


if __name__ == "__main__":
    row = find_row('Ароматизаторы.xlsx', 'Ароматизатор меловой SPIRIT REFILL - SAMURAI MAN')
    print(row)
    # 'L215444TE1'

